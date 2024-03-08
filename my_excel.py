from openpyxl import load_workbook, Workbook, cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string
from datetime import datetime
import csv
import pyuca
from cn2an import an2cn


class Excel_robot:
    """用于处理 Excel 数据的类"""

    def __init__(
        self,
        filename: str,
        data_only: bool,
        filename_summery: str | None = None,
        read_only: bool = False,
    ):
        """初始化 Excel_robot 对象"""
        self.filename = filename
        self.filename_summery = filename_summery
        self.wb = load_workbook(self.filename, data_only=data_only, read_only=read_only)
        self.max_row: int | None = None
        self.max_col: int | None = None
        self.summery_wb = load_workbook("./红白榜结果汇总.xlsx")

        # 获取红榜（男）、红榜（女）、较差、白榜工作表
        self.red_male_sheet = self.summery_wb["红榜（男）"]
        self.red_female_sheet = self.summery_wb["红榜（女）"]
        self.green_sheet = self.summery_wb["较差"]
        self.yellow_sheet = self.summery_wb["白榜"]
        self.red_male_start_row = 4
        self.red_female_start_row = 4
        self.green_start_row = 4
        self.yellow_start_row = 4

    def get_sheet_names(self, wb: Workbook):
        """
        打印所有的sheet名称
        """
        sheet_names = wb.sheetnames
        print("sheets: " + str(sheet_names), end="\n\n")
        return sheet_names

    def get_data(
        self,
        sheet_name: str,
        max_col_string: str,
        max_row: int | None = None,
        contains_first_line=False,
    ):
        """从指定的工作表中获取数据，并导出到 CSV 文件"""
        self.max_row = max_row
        self.max_col = column_index_from_string(max_col_string)
        ws = self.wb[sheet_name]

        data = []
        # 遍历所有cell
        for row in ws.iter_rows(
            min_row=2 if not contains_first_line else 1,
            max_col=self.max_col,
            max_row=self.max_row,
        ):
            data.append([cell.value for cell in row])

        with open(f"{sheet_name}.csv", "w", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(value for value in data)

        print(f"{sheet_name}.csv dump ok...\n")  # 提示完成
        return sheet_name

    def read_csv(self, filename: str):
        """从 CSV 文件中读取数据"""
        data = []
        with open(filename + ".csv", "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            # 处理data的None值, 如果为存在None则忽略当行
            for row in reader:
                if None in row or "" in row or len(row) < 1 or int(row[-1]) == 0:
                    continue
                # filtered_row = [value for value in row]
                # 将每行的最后一个字符串数字转换为整数
                last_element_int = int(row[-1])

                # 打印每行的内容（除了最后一个元素为整数外，其他元素仍然为字符串）
                filtered_row = row[:-1] + [last_element_int]
                data.append(filtered_row)

        return data

    def sorting_column(
        self,
        data: list,
        col_string: str,
        desc=False,
        convert_to_int=False,
    ):
        """根据指定列对数据进行排序"""
        sorted_col_index = column_index_from_string(col_string) - 1
        if convert_to_int:
            data.sort(key=lambda x: int(x[sorted_col_index]), reverse=desc)
        else:
            collator = pyuca.Collator()
            data.sort(
                key=lambda x: collator.sort_key(x[sorted_col_index]), reverse=True
            )

    def sorting_data(self, data) -> list:
        """对数据进行排序"""
        self.sorting_column(data, "B")
        self.sorting_column(data, "A")
        self.sorting_column(data, "E", desc=True, convert_to_int=True)
        return data

    def write_to_wb(
        self,
        sheet_name: str,
        sorted_data,
        start_row: int,
    ):
        """将排序后的数据写回到工作表中"""
        wb = self.wb
        ws = wb[sheet_name]
        count_row = len(sorted_data) + 1
        for i, row in enumerate(
            ws.iter_rows(
                min_row=start_row,
                max_row=self.max_row,
                max_col=self.max_col,
            )
        ):
            for j, cell in enumerate(row):
                try:
                    cell.value = sorted_data[i][j]
                    if j == column_index_from_string("E") - 1:
                        self.fill_color(cell=cell)

                except Exception as e:
                    pass

        for row in ws.iter_rows(
            min_row=count_row + 1, max_col=column_index_from_string("E")
        ):
            for cell in row:
                cell.value = None

    def divide_score(self, score: int):
        """将分数分为不同阶段"""
        if score >= 100:
            return "red"
        if 60 <= score <= 65:
            return "green"
        if 0 < score < 60:
            return "yellow"
        return None

    def fill_color(self, cell: cell):
        """根据分数不同阶段填充单元格颜色"""
        red = PatternFill("solid", fgColor="00FF0000")
        green = PatternFill("solid", fgColor="92D050")
        yellow = PatternFill("solid", fgColor="00FFFF00")

        value = int(cell.value)
        stage = self.divide_score(value)
        if stage == "red":
            cell.fill = red
        elif stage == "green":
            cell.fill = green
        elif stage == "yellow":
            cell.fill = yellow
        else:
            pass

    def summery(
        self,
        sorted_data,
        sheet_name,
    ):
        """将分类后的数据汇总到汇总表中"""
        for i, row in enumerate(sorted_data):

            if sheet_name == "女生排序":
                if i == 110:
                    print()

            score = int(row[-1])

            if None in set(row):
                continue

            stage = self.divide_score(score)
            if stage is None:
                continue

            if stage == "red":
                if sheet_name == "男生排序":
                    ws = self.red_male_sheet
                    self.red_male_start_row += 1
                    current_row = self.red_male_start_row
                if sheet_name == "女生排序":
                    ws = self.red_female_sheet
                    self.red_female_start_row += 1
                    current_row = self.red_female_start_row
            elif stage == "green":
                ws = self.green_sheet
                self.green_start_row += 1
                current_row = self.green_start_row
            elif stage == "yellow":
                ws = self.yellow_sheet
                self.yellow_start_row += 1
                current_row = self.yellow_start_row

            for j, value in enumerate(row):
                if j == len(row) - 1:
                    continue
                ws.cell(row=current_row, column=j + 1, value=value)

    def sign_content(self, ws: Worksheet, row, column, formatted_date_now):

        ws.cell(row=row, column=column, value="电子信息学院")
        ws.cell(row=row + 1, column=column, value=formatted_date_now)

        # sign_content(ws=self.red_male_sheet, row=red_male_end_row, column=D_index)
        # sign_content(self.red_female_sheet, row=red_female_end_row, column=D_index)
        # sign_content(ws=self.green_sheet, row=green_end_row, column=E_index)
        # sign_content(ws=self.yellow_sheet, row=yellow_end_row, column=E_index)


def calculate_date(start_date_string="2024-02-19"):
    """计算当前日期与开始日期之间的天数，以及当前是第几周"""
    date_obj = datetime.strptime(start_date_string, "%Y-%m-%d")
    current_date = datetime.now()
    formatted_date_now = current_date.strftime("%Y年%m月%d日")
    delta = current_date - date_obj
    day_num = delta.days
    count_week = int(day_num / 7)
    chinese_number = an2cn(count_week)
    week_string = "第" + chinese_number + "周"
    return week_string, formatted_date_now


if __name__ == "__main__":

    # 创建对象
    robot = Excel_robot(
        filename="./红白榜数据汇总.xlsx",
        # filename_summery="./红白榜结果汇总.xlsx",
        data_only=True,
    )
    # 获取所有sheet
    sheet_names = robot.get_sheet_names(robot.wb)

    # 导出sheet的内容到csv,并返回当前sheet名称
    sheet_male = robot.get_data(
        sheet_names[-1],
        max_col_string="E",
        max_row=None,
    )
    # print(f"{sheet_male}.csv dump ok...\n")

    sheet_female = robot.get_data(
        sheet_names[-2],
        max_col_string="E",
        max_row=None,
    )
    # print(f"{sheet_female}.csv dump ok...\n")

    # 得到处理过的data
    processed_male_data = robot.read_csv(sheet_male)
    processed_female_data = robot.read_csv(sheet_female)

    # 排序
    sorted_male_data = robot.sorting_data(processed_male_data)  # 男生排序
    sorted_female_data = robot.sorting_data(processed_female_data)  # 女生排序

    # 写入Excel并填充颜色
    robot.write_to_wb(sheet_male, sorted_data=sorted_male_data, start_row=2)
    robot.write_to_wb(sheet_female, sorted_data=sorted_female_data, start_row=2)
    robot.wb.save(robot.filename + "sorted.xlsx")
    print(f"{robot.filename} 已经保存\n")

    # 计算日期
    chinese_number, formatted_date_now = calculate_date()

    # 自动写入第几周
    robot.red_male_sheet["A2"] = chinese_number
    robot.red_female_sheet["A2"] = chinese_number
    robot.green_sheet["A2"] = chinese_number
    robot.yellow_sheet["A2"] = chinese_number

    # 将各个阶段的分数写入汇总表
    robot.summery(
        sorted_data=sorted_male_data,
        sheet_name=sheet_male,
    )
    robot.summery(
        sorted_data=sorted_female_data,
        sheet_name=sheet_female,
    )
    red_male_end_row = robot.red_male_start_row + 1
    red_female_end_row = robot.red_female_start_row + 1
    green_end_row = robot.green_start_row + 1
    yellow_end_row = robot.yellow_start_row + 1

    E_index = column_index_from_string("E")
    D_index = column_index_from_string("D")

    robot.sign_content(
        ws=robot.red_male_sheet,
        row=red_male_end_row,
        column=D_index,
        formatted_date_now=formatted_date_now,
    )
    robot.sign_content(
        robot.red_female_sheet,
        row=red_female_end_row,
        column=D_index,
        formatted_date_now=formatted_date_now,
    )
    robot.sign_content(
        ws=robot.green_sheet,
        row=green_end_row,
        column=E_index,
        formatted_date_now=formatted_date_now,
    )
    robot.sign_content(
        ws=robot.yellow_sheet,
        row=yellow_end_row,
        column=E_index,
        formatted_date_now=formatted_date_now,
    )

    robot.summery_wb.save("./汇总.xlsx")
    print("ok")
