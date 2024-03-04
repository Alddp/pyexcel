from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 加载红白榜数据汇总.xlsx文件
wb = load_workbook("./红白榜数据汇总.xlsx", data_only=True)
# options = {"": "", "": ""}
# 设置排序方式为男生排序
cho = "男生排序"

# 获取男生排序表的数据
ws = wb[cho]


# 获取男生排序表的数据
def get_data(ws):
    data = []

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=471)):
        data.append([])

        for cell in row:
            data[i].append(cell.value)
    return data


def sorting_score(data, desc=True):
    # 根据"E"分数排序
    sorted_data = sorted(data, key=lambda x: x[-1], reverse=desc)
    return sorted_data


def sorting_building_no(data, desc=False):
    # 根据"A"楼号排序
    sorted_building_no = sorted(data, key=lambda x: x[0], reverse=desc)
    return sorted_building_no


def sorting_dormitory_no(data, desc=False):
    # 根据"B"宿舍号排序
    sorted_dormitory_no = sorted(data, key=lambda x: x[1], reverse=desc)
    return sorted_dormitory_no


def write_sorted_data(sorted_data, ws):
    # 写入EXCEL
    for i, row in enumerate(sorted_data, start=1):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i + 1, column=j, value=value)


def main_sorting(data):
    # 主要排序:分数 降序
    # 次要排序:楼号, 宿舍号 升序
    # 先进行次要排序

    sorted_building_data = sorting_building_no(data)
    sorted_dormitory_data = sorting_dormitory_no(sorted_building_data)
    sorted_data = sorting_score(sorted_dormitory_data)
    return sorted_data


# 根据"E"分数填充颜色
def fill_color(ws):
    """
    score=100 红色
    score>60 & score<65 绿色
    score<60 黄色
    """
    fill_red = PatternFill("solid", fgColor="00FF0000")
    fill_green = PatternFill("solid", fgColor="92D050")
    fill_yellow = PatternFill("solid", fgColor="00FFFF00")

    for row in ws.iter_rows(min_row=2):
        score = int(row[-1].value)
        for cell in row:
            if score == 100:
                cell.fill = fill_red
            if 60 < score < 65:
                cell.fill = fill_green
            if 0 < score < 60:
                cell.fill = fill_yellow


if __name__ == "__main__":
    # 获取男生排序表的数据
    data = get_data(ws)
    # 主要排序:分数 降序
    sorted_data = main_sorting(data)
    # 将排序后的数据写入EXCEL
    write_sorted_data(sorted_data, ws)
    # 根据"E"分数填充颜色
    fill_color(ws)

    wb.save("./sorted.xlsx")
    print("OK")
