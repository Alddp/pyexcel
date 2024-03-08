from openpyxl import load_workbook, Workbook, cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string
import csv
import pyuca
from summery import Summery


class Excel_robot:
    """docstring for Excel_robot."""

    def __init__(
        self,
        filename: str,
        data_only: bool,
        filename_summery: str = str | None,
        read_only: bool = False,
    ):
        self.filename = filename
        self.filename_summery = filename_summery
        self.wb = load_workbook(self.filename, data_only=data_only, read_only=read_only)
        self.max_row: int | None = None
        self.max_col: int | None = None

    def get_sheet_names(self, wb: Workbook):
        """
        打印所有的sheet名称
        """
        sheet_names = wb.sheetnames
        print("sheets: " + str(sheet_names), end="\n\n")
        return sheet_names

    def get_data(
        self,
        sheet_name: str,
        max_col_string: str,
        max_row: int | None = None,
        contains_first_line=False,
    ):
        """
        读取sheet所有内容写入data
        sheet_name: 要读取的列表
        """
        self.max_row = max_row
        self.max_col = column_index_from_string(max_col_string)
        ws = self.wb[sheet_name]

        data = []
        # 遍历所有cell
        for row in ws.iter_rows(
            min_row=2 if not contains_first_line else 1,
            max_col=self.max_col,
            max_row=self.max_row,
        ):
            data.append([cell.value for cell in row])

        with open(f"{sheet_name}.csv", "w", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(value for value in data)

        return sheet_name

    def read_csv(self, filename: str):
        data = []
        with open(filename + ".csv", "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            # 处理data的None值, 如果为存在None则忽略当行
            for row in reader:
                # print(row)
                if None in row or "" in row or len(row) < 1:
                    continue
                filtered_row = [value for value in row]
                data.append(filtered_row)

        return data

    def sorting_column(
        self,
        data: list,
        col_string: str,
        desc=False,
        convert_to_int=False,
    ):
        """
        根据所选列的值排序
        """

        # 创建一个Unicode排序器
        collator = pyuca.Collator()

        # sorted_col_index = column_index_from_string(col_string) - 1
        # sorted_data = sorted(data, key=lambda x: x[sorted_col_index], reverse=desc)
        # return sorted_data

        sorted_col_index = column_index_from_string(col_string) - 1
        if convert_to_int:
            data.sort(key=lambda x: int(x[sorted_col_index]), reverse=desc)
        else:
            # data.sort(key=lambda x: x[sorted_col_index], reverse=desc)
            data.sort(
                key=lambda x: collator.sort_key(x[sorted_col_index]), reverse=True
            )

    def sorting_data(self, data):
        """
        总排序方法
        """
        self.sorting_column(data, "B")
        self.sorting_column(data, "A")
        self.sorting_column(data, "E", desc=True, convert_to_int=True)
        return data

    # TODO: 将排序好的data写入excel
    def write_to_wb(
        self,
        sheet_name: str,
        sorted_data,
        start_row: int,
        summery_filename: str,
    ):
        wb = self.wb
        ws = wb[sheet_name]

        for i, row in enumerate(
            ws.iter_rows(
                min_row=start_row,
                max_row=self.max_row + 1,
                max_col=self.max_col,
            )
        ):
            for j, cell in enumerate(row):
                try:
                    cell.value = sorted_data[i][j]
                except Exception as e:
                    pass
