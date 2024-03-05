# 对比两异
from openpyxl import load_workbook
import csv


def get_csv(sheet_name, filename):
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name]
    with open(filename + ".csv", "w", encoding="utf-8") as f:
        writer = csv.writer(f)
        row_title = []
        # 第二列
        for row in ws.iter_rows(min_col=1, max_col=5):
            row_title.append(cell.value for cell in row)
        writer.writerows(row_title)


if __name__ == "__main__":

    sheet_name = "男生排序"
    get_csv(sheet_name, filename="红白榜数据汇总.xlsxsorted.xlsx")
    get_csv(sheet_name, filename="对比.xlsx")

    print("OK")
